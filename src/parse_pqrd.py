"""
parse_pqrd.py
Extrae las tablas de reclamos en salud de los reportes Excel "año corrido"
de Supersalud (data/raw/*.xlsx) y las consolida en un CSV limpio.

IMPORTANTE: las tablas NO están en la misma fila exacta entre archivos de
distintos años, por eso este script busca cada tabla POR TEXTO (el nombre
de la fila/etiqueta), nunca por número de fila fijo.

Uso:
    python src/parse_pqrd.py
"""

import glob
import os
import re
import openpyxl
import pandas as pd

RUTA_RAW = "data/raw"
RUTA_PROCESSED = "data/processed"
MESES = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN",
         "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]

# Etiquetas que buscamos dentro de cada archivo. Si Supersalud cambia el
# texto exacto en el futuro, ajustar aquí (con variantes en la lista).
#
# HALLAZGO IMPORTANTE (confirmado inspeccionando los archivos crudos):
# Supersalud cambió la taxonomía de "motivos específicos" y la de "tipo de
# PQRD" ENTRE 2023 Y 2024 (no entre 2021 y 2022 como se pensó inicialmente).
# - 2021, 2022, 2023: tabla "TIPO DE PQRD" (categorías REGULARES/SIS) y
#   10 motivos específicos orientados a especialidades médicas y trámites
#   clínicos (ver MOTIVO_ESPECIFICO_LEGACY).
# - 2024, 2025, 2026: tabla "TIPO DE RIESGO" (SIMPLE/PRIORIZADO/RIESGO
#   VITAL) y 10 motivos específicos orientados a barreras de acceso y
#   autorizaciones (ver MOTIVO_ESPECIFICO_NUEVO).
# Ambas taxonomías se extraen y se guardan CON ETIQUETAS DE ORIGEN
# DISTINTAS (tabla_origen = 'motivo_especifico_2024_2026' vs
# 'motivo_especifico_legacy_2021_2023') para que NUNCA se mezclen en el
# análisis como si fueran la misma categoría.

MOTIVO_ESPECIFICO_NUEVO = [
    "NEGACIÓN PARA LA ENTREGA DE TECNOLOGÍAS EN SALUD Y/O DE OTROS SERVICIOS AUTORIZADOS",
    "NEGACIÓN EN LA ASIGNACIÓN DE CITAS O CONSULTAS",
    "FALTA DE OPORTUNIDAD EN LAS CITAS O CONSULTAS",
    "FALTA DE OPORTUNIDAD EN LA ATENCIÓN EN OTROS SERVICIOS DE SALUD",
    "FALTA DE OPORTUNIDAD EN LA ENTREGA O ENTREGA INCOMPLETA DE TECNOLOGÍAS EN SALUD Y/O PRESTACIÓN DE OTROS SERVICIOS",
    "NEGACIÓN EN LA ATENCIÓN EN OTROS SERVICIOS DE SALUD",
    "FALTA DE OPORTUNIDAD EN LA AUTORIZACIÓN DE OTROS SERVICIOS DE SALUD",
    "FALTA DE OPORTUNIDAD EN LA AUTORIZACIÓN DE TECNOLOGÍAS EN SALUD Y/O DE OTROS SERVICIOS",
    "FALTA DE OPORTUNIDAD EN LA AUTORIZACIÓN DE CITAS DE CONSULTA",
    "FALTA DE OPORTUNIDAD EN EL PROCESO DE REFERENCIA Y CONTRARREFERENCIA",
]

MOTIVO_ESPECIFICO_LEGACY = [
    "FALTA DE OPORTUNIDAD EN LA ASIGNACIÓN DE CITAS DE CONSULTA MÉDICA ESPECIALIZADA",
    "FALTA DE OPORTUNIDAD EN LA ENTREGA DE MEDICAMENTOS POS",
    "DEMORA DE LA PROGRAMACIÓN DE EXÁMENES DE LABORATORIO O DIAGNÓSTICOS",
    "FALTA DE OPORTUNIDAD EN LA PROGRAMACIÓN DE CIRUGÍA",
    "FALTA DE OPORTUNIDAD PARA LA PRESTACIÓN DE SERVICIOS DE IMAGENOLOGÍA",
    "NO APLICACIÓN DE NORMAS, GUÍAS O PROTOCOLOS DE ATENCIÓN",
    "DEFICIENCIAS EN LA SEGURIDAD DEL PACIENTE",
    "FALTA DE OPORTUNIDAD EN LA ENTREGA DE MEDICAMENTOS NO POS",
    "FALTA DE OPORTUNIDAD EN LA ASIGNACIÓN DE CITAS DE CONSULTA MÉDICA GENERAL",
    "DEFICIENTE INFORMACIÓN SOBRE DERECHOS, DEBERES Y TRÁMITES",
]

ETIQUETAS_TABLAS = {
    "capital_departamento": ["MEDELLÍN", "MEDELLIN"],
    "motivo_especifico_2024_2026": MOTIVO_ESPECIFICO_NUEVO,
    "motivo_especifico_legacy_2021_2023": MOTIVO_ESPECIFICO_LEGACY,
    "departamento": ["ANTIOQUIA"],
    "otro_vigilado": ["DIRECCION SECCIONAL DE SALUD DE ANTIOQUIA",
                       "DIRECCIÓN SECCIONAL DE SALUD DE ANTIOQUIA"],
}


def extraer_anio_del_nombre(nombre_archivo: str) -> int | None:
    """Busca un año de 4 dígitos (20XX) en el nombre del archivo."""
    match = re.search(r"20\d{2}", nombre_archivo)
    return int(match.group()) if match else None


def normalizar_texto(valor) -> str:
    """Uppercase y quita tildes, para comparar sin importar acentos."""
    if valor is None:
        return ""
    import unicodedata
    texto = str(valor).strip().upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto


def encontrar_hoja_principal(wb: openpyxl.Workbook) -> str:
    """Devuelve el nombre de la hoja que contiene los datos ('TASA AÑO CORRIDO'
    en archivos 2022+, 'PQRD - SNS' en el archivo 2021)."""
    candidatos = ["TASA AÑO CORRIDO", "PQRD - SNS"]
    for nombre in wb.sheetnames:
        if normalizar_texto(nombre) in [normalizar_texto(c) for c in candidatos]:
            return nombre
    # si no encuentra ninguno conocido, usa la primera hoja
    return wb.sheetnames[0]


def buscar_fila_por_texto(ws, texto_buscado: str, columnas_a_revisar: range = range(1, 8)) -> int | None:
    """Busca el texto en CUALQUIERA de las primeras columnas (A-G) de cada
    fila, usando ws.cell(row, column) directamente -- NO iter_rows(), porque
    iter_rows() alinea los índices al rango usado de la hoja (min_col), lo
    cual desalinea la búsqueda si la hoja no empieza en la columna A.

    IMPORTANTE (bug corregido): solo se compara en UNA dirección
    (texto_buscado dentro del valor de la celda). Comparar en ambas
    direcciones causaba falsos positivos graves: textos cortos como "ENE"
    (valor de columna de mes) "cabían" dentro de textos largos como
    "DEMORA DE LA PROGRAMACIÓN DE EXÁMENES..." por pura coincidencia de
    letras (ENE está contenido en EXÁMENES), sin relación real con el
    contenido buscado. Con una sola dirección, Python ya descarta
    automáticamente los casos donde el valor de la celda es más corto que
    el texto buscado (in() nunca es True si la cadena buscada es más larga
    que donde se busca), así que no hace falta un filtro de longitud aparte.

    Devuelve el número de fila o None si no se encuentra."""
    texto_buscado_norm = normalizar_texto(texto_buscado)
    max_fila = ws.max_row

    for fila in range(1, max_fila + 1):
        for columna in columnas_a_revisar:
            valor = ws.cell(row=fila, column=columna).value
            if valor is None:
                continue
            valor_norm = normalizar_texto(valor)
            if not valor_norm:
                continue
            if texto_buscado_norm in valor_norm:
                return fila
    return None


def leer_valores_mensuales(ws, fila: int, col_inicio: int = 4) -> dict:
    """Lee los 12 valores mensuales (ENE-DIC) a partir de la columna dada
    en la fila indicada. Devuelve dict {mes: valor}, con None para meses
    sin datos (no confundir con 0)."""
    valores = {}
    for i, mes in enumerate(MESES):
        celda = ws.cell(row=fila, column=col_inicio + i)
        valores[mes] = celda.value  # puede ser None; NO forzar a 0
    return valores


def procesar_archivo(ruta_archivo: str) -> list:
    """Procesa un archivo Excel y devuelve una lista de registros
    (uno por categoría x mes) en formato largo (tidy)."""
    nombre_archivo = os.path.basename(ruta_archivo)
    anio = extraer_anio_del_nombre(nombre_archivo)
    if anio is None:
        print(f"[WARN] No se pudo determinar el año de '{nombre_archivo}'. Se omite.")
        return []

    print(f"[INFO] Procesando {nombre_archivo} (año detectado: {anio})...")
    wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
    hoja = encontrar_hoja_principal(wb)
    ws = wb[hoja]

    registros = []

    # --- Medellín (tabla "por capital de departamento") ---
    fila_medellin = buscar_fila_por_texto(ws, "MEDELLÍN")
    if fila_medellin:
        valores = leer_valores_mensuales(ws, fila_medellin)
        for mes, valor in valores.items():
            if valor is not None:
                registros.append({
                    "anio": anio, "mes": mes, "tabla_origen": "capital_departamento",
                    "categoria": "MEDELLIN", "valor": valor,
                })
    else:
        print(f"[WARN] No se encontró la fila de Medellín en {nombre_archivo}.")

    # --- Antioquia (tabla "por departamento") ---
    fila_antioquia = buscar_fila_por_texto(ws, "ANTIOQUIA")
    if fila_antioquia:
        valores = leer_valores_mensuales(ws, fila_antioquia)
        for mes, valor in valores.items():
            if valor is not None:
                registros.append({
                    "anio": anio, "mes": mes, "tabla_origen": "departamento",
                    "categoria": "ANTIOQUIA", "valor": valor,
                })

    # --- Motivos específicos: taxonomía NUEVA (2024-2026) ---
    for motivo in MOTIVO_ESPECIFICO_NUEVO:
        fila_motivo = buscar_fila_por_texto(ws, motivo)
        if fila_motivo:
            valores = leer_valores_mensuales(ws, fila_motivo)
            for mes, valor in valores.items():
                if valor is not None:
                    registros.append({
                        "anio": anio, "mes": mes,
                        "tabla_origen": "motivo_especifico_2024_2026",
                        "categoria": motivo[:60], "valor": valor,
                    })

    # --- Motivos específicos: taxonomía LEGACY (2021-2023) ---
    for motivo in MOTIVO_ESPECIFICO_LEGACY:
        fila_motivo = buscar_fila_por_texto(ws, motivo)
        if fila_motivo:
            valores = leer_valores_mensuales(ws, fila_motivo)
            for mes, valor in valores.items():
                if valor is not None:
                    registros.append({
                        "anio": anio, "mes": mes,
                        "tabla_origen": "motivo_especifico_legacy_2021_2023",
                        "categoria": motivo[:60], "valor": valor,
                    })

    # --- Dirección Seccional de Salud de Antioquia (tabla "otro vigilado") ---
    for etiqueta in ETIQUETAS_TABLAS["otro_vigilado"]:
        fila = buscar_fila_por_texto(ws, etiqueta)
        if fila:
            valores = leer_valores_mensuales(ws, fila)
            for mes, valor in valores.items():
                if valor is not None:
                    registros.append({
                        "anio": anio, "mes": mes, "tabla_origen": "otro_vigilado",
                        "categoria": "DIRECCION_SECCIONAL_SALUD_ANTIOQUIA", "valor": valor,
                    })
            break

    print(f"       -> {len(registros)} registros extraídos de {nombre_archivo}")
    return registros


def main():
    archivos = sorted(glob.glob(os.path.join(RUTA_RAW, "*.xlsx")))
    if not archivos:
        print(f"[ERROR] No se encontraron archivos .xlsx en {RUTA_RAW}/. "
              f"Verifica que subiste los 6 archivos ahí.")
        return

    todos_los_registros = []
    for archivo in archivos:
        registros = procesar_archivo(archivo)
        todos_los_registros.extend(registros)

    if not todos_los_registros:
        print("[ERROR] No se extrajo ningún registro de ningún archivo. "
              "Revisa los nombres de hoja/etiquetas.")
        return

    df = pd.DataFrame(todos_los_registros)

    # Nota metodológica: años <2022 pueden tener menos registros porque
    # la clasificación de "motivo específico" no existía con ese detalle.
    os.makedirs(RUTA_PROCESSED, exist_ok=True)
    ruta_salida = os.path.join(RUTA_PROCESSED, "pqrd_consolidado.csv")
    df.to_csv(ruta_salida, index=False, encoding="utf-8")

    print("\n" + "=" * 60)
    print(f"[OK] Consolidado guardado en {ruta_salida}")
    print(f"     Total registros: {len(df)}")
    print(f"     Años cubiertos: {sorted(df['anio'].unique())}")
    print(f"     Tablas de origen: {df['tabla_origen'].unique().tolist()}")
    print("=" * 60)
    print("\nNOTA METODOLÓGICA (verificar que esto se cumpla, documentar en")
    print("docs/marco_metodologico.md si hay desviaciones):")
    print("  - 'motivo_especifico_2024_2026' debe tener registros SOLO en 2024, 2025, 2026")
    print("  - 'motivo_especifico_legacy_2021_2023' debe tener registros SOLO en 2021, 2022, 2023")
    print("  - 'capital_departamento', 'departamento', 'otro_vigilado' deben")
    print("    tener registros en TODOS los años (2021-2026), porque esas")
    print("    categorías (nombres de ciudades/departamentos/entidades) no")
    print("    cambiaron de taxonomía en el tiempo.")
    print()
    print(df.groupby(["anio", "tabla_origen"]).size().unstack(fill_value=0))


if __name__ == "__main__":
    main()